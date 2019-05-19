# -----------------------------------------------------------------------------
# Copyright (c) 2015, Nicolas P. Rougier. All Rights Reserved.
# Distributed under the (new) BSD License. See LICENSE.txt for more info.
# -----------------------------------------------------------------------------
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.cbook as cbook
from openpyxl import *
from pylab import *
import datetime
# from openpyxl import load_workbook
from pylab import *

mpl.rcParams['font.sans-serif'] = ['SimHei']

# ---------------------------------------------------------------------
#   从文本中提取数据
#   数据关键字：时间X、站点Y
#
# ---------------------------------------------------------------------

data_wb = load_workbook('分析后数据表.xlsx')
data_ws = data_wb['Sheet']

# ---------------------------------------------------------------------
#   使用散点图显示车站与时间之间的关系
#   数据关键字：站点X、时间Y、故障数量s
#
# ---------------------------------------------------------------------
'''
# 建立车站列表，自东莞火车站虎门火车站（0-14），车辆段15，控制中心16，旗峰主所17，厚街主所18，区间19
datad_station = {'dgz': '0', 'csz': '0', 'lhz': '0', 'xqz': '0', 'tbz': '0', 'dcz': '0', 'qfz': '0', 'hfz': '0',
                 'xpz': '0', 'hdz': '0', 'cwz': '0', 'lxz': '0', 'smz': '0', 'zlz': '0', 'hmz': '0', 'socc': '0',
                 'occ': '0', 'qfzs': '0', 'hjzs': '0', 'qj': '0', }
for data_ws_i in range(data_ws.max_row):
    if data_ws.cell(data_ws_i + 1, 7).value == '东莞站':
        datad_station['dgz'] = int(datad_station['dgz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '茶山站':
        datad_station['csz'] = int(datad_station['csz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '榴花站':
        datad_station['lhz'] = int(datad_station['lhz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '下桥站':
        datad_station['xqz'] = int(datad_station['xqz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '天宝站':
        datad_station['tbz'] = int(datad_station['tbz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '东城站':
        datad_station['dcz'] = int(datad_station['dcz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '旗峰站':
        datad_station['qfz'] = int(datad_station['qfz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '鸿福站':
        datad_station['hfz'] = int(datad_station['hfz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '西平站':
        datad_station['xpz'] = int(datad_station['xpz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '蛤地站':
        datad_station['hdz'] = int(datad_station['hdz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '陈屋站':
        datad_station['cwz'] = int(datad_station['cwz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '寮夏站':
        datad_station['lxz'] = int(datad_station['lxz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '珊美站':
        datad_station['smz'] = int(datad_station['smz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '展览站':
        datad_station['zlz'] = int(datad_station['zlz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '虎门站':
        datad_station['hmz'] = int(datad_station['hmz']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '车辆段':
        datad_station['socc'] = int(datad_station['socc']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '西平控制中心':
        datad_station['occ'] = int(datad_station['occ']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '旗峰主所':
        datad_station['qfzs'] = int(datad_station['qfzs']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '厚街主所':
        datad_station['hjzs'] = int(datad_station['hjzs']) + 1
    elif data_ws.cell(data_ws_i + 1, 7).value == '区间':
        datad_station['qj'] = int(datad_station['qj']) + 1

Y = ('东莞站','茶山站','榴花站','下桥站','天宝站','东城站','旗峰站','鸿福路站','西平站','蛤地站','陈屋站','寮夏站','珊美站','展览站','虎门站','车辆段','西平控制中心','旗峰主所','厚街主所','区间')
X = np.random.rand(20)

area = list(datad_station.values())

print(area)
k=0
fig = plt.figure()
ax = plt.subplot()
for k in range(20):
    ax.scatter(X, Y, s=area[k])


#scatter(X, Y, s=area, alpha = 0.6)
plt.show()
'''

# ---------------------------------------------------------------------
#   使用折现图显示专业故障总数与时间之间的关系
#   数据关键字：时间X、单位时间内故障总数Y
#   data_ws_k为顺序统计单元格4中日期顺序临时变量
# ---------------------------------------------------------------------
date = {'一月': 0, '二月': 0, '三月': 0, '四月': 0, '五月': 0, '六月': 0, '七月': 0, '八月': 0, '九月': 0, '十月': 0, '十一月': 0, '十二月': 0}

for data_ws_k in range(data_ws.max_row):
    if datetime.datetime(2019, 1, 1) <= data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2019, 2, 1):
        date['一月'] = int(date['一月']) + 1
    elif datetime.datetime(2019, 2, 1) < data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2019, 3, 1):
        date['二月'] = int(date['二月']) + 1
    elif datetime.datetime(2019, 3, 1) < data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2019, 4, 1):
        date['三月'] = int(date['三月']) + 1
    elif datetime.datetime(2019, 4, 1) < data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2019, 5, 1):
        date['四月'] = int(date['四月']) + 1
    elif datetime.datetime(2019, 5, 1) < data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2019, 6, 1):
        date['五月'] = int(date['五月']) + 1
    elif datetime.datetime(2019, 6, 1) < data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2019, 7, 1):
        date['六月'] = int(date['六月']) + 1
    elif datetime.datetime(2019, 7, 1) < data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2019, 8, 1):
        date['七月'] = int(date['七月']) + 1
    elif datetime.datetime(2019, 8, 1) < data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2019, 9, 1):
        date['八月'] = int(date['八月']) + 1
    elif datetime.datetime(2019, 9, 1) < data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2019, 10, 1):
        date['九月'] = int(date['九月']) + 1
    elif datetime.datetime(2019, 10, 1) < data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2019, 11, 1):
        date['十月'] = int(date['十月']) + 1
    elif datetime.datetime(2019, 11, 1) < data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2019, 12, 1):
        date['十一月'] = int(date['十一月']) + 1
    elif datetime.datetime(2019, 12, 1) < data_ws.cell(data_ws_k + 1, 4).value < datetime.datetime(2020, 1, 1):
        date['十二月'] = int(date['十二月']) + 1

figure()
xlabel("月份")
ylabel("月度故障总数")
title("月度故障总数折线图")
Y = list(date.values())
X = ('一月', '二月', '三月', '四月', '五月', '六月', '七月', '八月', '九月', '十月', '十一月', '十二月')
plot(X, Y)
show()
