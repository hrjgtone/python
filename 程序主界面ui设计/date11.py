#!/usr/bin/env python3.7
#-*- coding:  utf-8 -*-
'a test project'
__author__='GTone'

from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
import tkinter.filedialog
from tkinter import ttk
import tkinter as tk



#---------------------------------------------------------------------
#   数据提取及处理类
#
#
#---------------------------------------------------------------------


#调取UI界面选择的xlsx文件数据并将综合监控专业相关数据列出保存至“分析后数据表”内
def read_sheet(filename):
    #新建表格
    wb2=Workbook()
    wb2.save('分析后数据表.xlsx')
    #定义所选文件名
    wb1 = load_workbook(filename)
    wb2 = load_workbook('分析后数据表.xlsx')
    #定义所选表格中的sheel
    ws1 = wb1['故障、问题统计子表']
    ws2 = wb2['Sheet']
    k=0
    #循环输出符合综合监控专业的故障
    for ws1_i in range(ws1.max_row):
        if ws1.cell(ws1_i+1,3).value=='综合监控'or ws1.cell(ws1_i+1,3).value=='PSCADA':
            k=k+1
            for j in range(ws1.max_column):
                ws2.cell(k,j+1).value=ws1.cell(ws1_i+1,j+1).value
    wb2.save('分析后数据表.xlsx')



#---------------------------------------------------------------------
#   UI界面类
#
#
#---------------------------------------------------------------------


#主界面UI设计
def main_ui():
    root = Tk()
    root.title('故障分析系统')
    root.geometry('640x320')
    ttk.Label(root,text="选择专业:").grid(column=1,row=0)
    
 
    #设置专业复选框
    global chVarOne
    global chVarTwo
    global chVarThree
    
    chVarOne = tk.IntVar()
    check1 = tk.Checkbutton(root, text="综合监控",variable=chVarOne,offvalue=0,onvalue=1,command=selected_major)
    check1.deselect()
    check1.grid(column=0, row=4, sticky=tk.W)

    chVarTwo = tk.IntVar()
    check2 = tk.Checkbutton(root, text="电力监控",variable=chVarTwo,offvalue=0,onvalue=1,command=selected_major)
    check2.deselect()
    check2.grid(column=1, row=4, sticky=tk.W)

    chVarThree = tk.IntVar()
    check3 = tk.Checkbutton(root, text="其他",variable=chVarThree,offvalue=0,onvalue=1,command=selected_major)
    check3.deselect()
    check3.grid(column=2, row=4, sticky=tk.W)



    #设置按钮并调用函数
    btn=Button(root,text = '选择文件',command=ui_send_filepath).grid(row=320,column=640)
    #btn.pack()
    root.mainloop()
#---------------------------------------------------------------------
#   后台运算类
#
#
#---------------------------------------------------------------------    
#传递专业复选项的值并作出反馈
def selected_major():
    if(chVarOne.get()==1):
        major="综合监控"
        #tx.insert(0,"综合监控")
    elif(chVarTwo.get()==1):
        major="电力监控"
    elif(chVarThree.get()==1):
        major="其他"
    elif(chVarOne.get()==1 and chVarTwo.get()==1):
        major="综合监控、电力监控"
    elif(chVarOne.get()==1 and chVarThree.get()==1):
        major="综合监控、其他"
    elif(chVarThree.get()==1 and chVarTwo.get()==1):
        major="电力监控、其他"
    elif(chVarOne.get()==1 and chVarTwo.get()==1 and chVarThree.get()==1):
        major="综合监控、电力监控、其他"
    return major
#提取所选文件的绝对路径
def ui_send_filepath():
    filename=tkinter.filedialog.askopenfilename()
    read_sheet(filename)
    
#def ui_confirm_windows():
    

#---------------------------------------------------------------------
#   主函数调用
#
#
#---------------------------------------------------------------------
main_ui()
